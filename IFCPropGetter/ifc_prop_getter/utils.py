# -*- coding: utf-8 -*-

"""工具函数模块"""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ifc_prop_getter.constants import DATE_FORMAT, INVALID_FILENAME_CHARS


def format_timestamp():
    """返回当前时间的 HH:MM:SS 格式字符串"""
    return datetime.now().strftime("%H:%M:%S")


def safe_str(value):
    """将任意值安全转换为字符串"""
    if value is None:
        return "N/A"
    return str(value)


def clean_filename(name):
    """替换文件名中的非法字符为下划线"""
    return INVALID_FILENAME_CHARS.sub('_', name)


def get_default_output_dir():
    """获取默认输出目录（桌面或用户目录）"""
    desktop = Path.home() / "Desktop"
    return str(desktop if desktop.exists() else Path.home())


def make_output_filename(base, ext):
    """生成带时间戳的输出文件名"""
    date_str = datetime.now().strftime(DATE_FORMAT)
    safe_base = clean_filename(base.strip() or "ifc_properties_export")
    return f"{safe_base}_{date_str}.{ext.lstrip('.')}"


def export_to_excel_with_format(df, filepath):
    """将 DataFrame 导出为带样式的 Excel 文件"""
    # 写入数据
    df.to_excel(filepath, index=False, sheet_name='Sheet1')

    # 加载工作簿处理样式
    wb = load_workbook(filepath)
    ws = wb['Sheet1']

    # 预定义样式对象
    thin_side = Side(style='thin')
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    header_font = Font(name='Times New Roman', size=12, bold=True)
    content_font = Font(name='Times New Roman', size=11, bold=False)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    max_row = ws.max_row
    max_col = ws.max_column

    # 批量应用样式
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = full_border
            cell.alignment = center_align

            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.font = content_font

    # 设置列宽与行高
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        if col_name == "GlobalId":
            ws.column_dimensions[col_letter].width = 32
        else:
            ws.column_dimensions[col_letter].width = 24

    ws.row_dimensions[1].height = 32
    wb.save(filepath)


def extract_property_from_psets(psets, prop_name):
    """从属性集字典中提取指定属性值"""
    if '.' in prop_name:
        pset_name, p_name = prop_name.split('.', 1)
        props = psets.get(pset_name)
        if props:
            return safe_str(props.get(p_name))
        return "N/A"
    else:
        for props in psets.values():
            if prop_name in props:
                return safe_str(props[prop_name])
        return "N/A"